import pandas as pd
import openpyxl
from PIL  import Image
import glob
from os import listdir

wb = openpyxl.load_workbook("....xlsx")
ws = wb.worksheets[0]

image_dir='C:/...../images'

for image in glob.glob(image_dir+'/*.jpg'):
    img=Image.open(image)
    
for image in glob.glob(image_dir+'/*.jpg'):
    img=Image.open(image)
    height=350
    width= int(height*(img.size[1]/img.size[0]))
    img=img.resize((height,width),Image.ANTIALIAS)
    img = img.rotate(270, Image.NEAREST, expand = 1) 
    img.save(image+'_resized.jpg')

resize_dir='C:/.../images/resized/'


for i in range(len(df)):
    ws['C6'].value=df['SignID'].iloc[i]
    ws['F6'].value=df["Street"].iloc[i]
    ws['F9'].value=df["Address"].iloc[i]
    ws['F6'].value=df["Street"].iloc[i]
    ws['C10'].value=df["MUTCDclass"].iloc[i]
    ws['C11'].value=df["MUTCDcode"].iloc[i]
    ws['F12'].value=df["SignCondition"].iloc[i]
    ws['F13'].value=df["SignText"].iloc[i]
    ws['F11'].value=df["LegendColor"].iloc[i]
    ws['F10'].value=df["BackColor"].iloc[i]
    file_suffix=str(df["SignID"].iloc[i])

    wb.save('9860_'+file_suffix+'.xlsx')
    
    
wb = openpyxl.load_workbook("Sign_Inventory2.xlsx")
ws = wb.worksheets[0]

imagesList = listdir(resize_dir)

for img in imagesList:
    img=Image.open(img)
    
for resize in glob.glob(resize_dir+'/*.jpg'):    
    img = openpyxl.drawing.image.Image(resize[0])
    img.anchor = 'I9'
    ws.add_image(img)